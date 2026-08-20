"""
Attendance Data Exporter Module for AI-Enabled Smart Attendance System.
Implements standard CSV export and styled Excel (.xlsx) export via OpenPyXL.
EXPORTS ONLY ACADEMIC ATTENDANCE METADATA. ZERO BIOMETRIC DATA EXPORTED.
"""

import csv
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def export_attendance_csv(
    records: List[Dict[str, Any]],
    output_path: Union[str, Path],
) -> Path:
    """
    Export filtered attendance records to a standard CSV file.

    :param records: List of formatted attendance record dicts.
    :param output_path: Output CSV file path.
    :return: Path object to written CSV file.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "Attendance ID",
        "Student ID",
        "Student Name",
        "Class-Section",
        "Date",
        "Time",
        "Status",
        "Recognition Method",
    ]

    with open(out, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)

        for rec in records:
            writer.writerow([
                rec.get("attendance_id", ""),
                rec.get("student_code", ""),
                rec.get("student_name", ""),
                rec.get("class_section", ""),
                rec.get("attendance_date", ""),
                rec.get("attendance_time", ""),
                rec.get("status", ""),
                rec.get("recognition_method", ""),
            ])

    logger.info(f"Attendance CSV exported successfully to: {out} ({len(records)} records)")
    return out


def export_attendance_excel(
    records: List[Dict[str, Any]],
    output_path: Union[str, Path],
    summary_data: Optional[List[Dict[str, Any]]] = None,
) -> Path:
    """
    Export filtered attendance records and student summary metrics to a styled Excel (.xlsx) workbook using OpenPyXL.

    :param records: List of formatted attendance record dicts.
    :param output_path: Output .xlsx file path.
    :param summary_data: Optional list of student summary analytics dicts.
    :return: Path object to written Excel workbook file.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # ------------------------------------------------------------------------
    # SHEET 1: Attendance Records
    # ------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Attendance Records"

    headers1 = [
        "Attendance ID",
        "Student ID",
        "Student Name",
        "Class-Section",
        "Date",
        "Time",
        "Status",
        "Method",
    ]
    ws1.append(headers1)

    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for rec in records:
        row = [
            rec.get("attendance_id", ""),
            rec.get("student_code", ""),
            rec.get("student_name", ""),
            rec.get("class_section", ""),
            rec.get("attendance_date", ""),
            rec.get("attendance_time", ""),
            rec.get("status", ""),
            rec.get("recognition_method", ""),
        ]
        ws1.append(row)

    # Format data rows
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=len(headers1)):
        for cell in row:
            cell.border = thin_border
            if cell.column in (1, 2, 4, 5, 6, 7, 8):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-adjust column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ------------------------------------------------------------------------
    # SHEET 2: Student Summary
    # ------------------------------------------------------------------------
    if summary_data is not None:
        ws2 = wb.create_sheet(title="Student Summary")
        headers2 = [
            "Student ID",
            "Student Name",
            "Class-Section",
            "Total Days",
            "Present Count",
            "Absent Count",
            "Late Count",
            "Attendance %",
        ]
        ws2.append(headers2)

        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for item in summary_data:
            ws2.append([
                item.get("student_code", ""),
                item.get("student_name", ""),
                item.get("class_section", ""),
                item.get("total_days", 0),
                item.get("present_count", 0),
                item.get("absent_count", 0),
                item.get("late_count", 0),
                f"{item.get('attendance_percentage', 0.0)}%",
            ])

        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(headers2)):
            for cell in row:
                cell.border = thin_border
                if cell.column in (1, 3, 4, 5, 6, 7, 8):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(out)
    logger.info(f"Attendance Excel workbook saved successfully to: {out}")
    return out
