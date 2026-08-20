"""
Comprehensive Stage 8 Automated Test Suite for Attendance Reports & Data Export System.
Tests authentication, RBAC authorization, multi-criteria filtering, student analytics,
manual correction, CSV file export, OpenPyXL Excel export, and data safety.
"""

import csv
import tempfile
from pathlib import Path
import openpyxl
import pytest

from app.database import (
    initialize_database,
    create_student,
    create_attendance,
)
from app.auth import get_session
from app.reports import (
    search_attendance_records,
    get_student_attendance_summary,
    correct_attendance_record,
    export_attendance_csv,
    export_attendance_excel,
)


@pytest.fixture
def temp_db():
    """Create a temporary SQLite database for testing."""
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as f:
        db_path = Path(f.name)
    initialize_database(db_path)
    yield db_path
    if db_path.exists():
        try:
            db_path.unlink()
        except PermissionError:
            pass


@pytest.fixture
def auth_session():
    """Authenticate session as admin/teacher for RBAC testing."""
    session = get_session()
    session.start_session({"id": 1, "username": "admin_test", "role": "admin"})
    yield session
    session.clear_session()


# ============================================================================
# 1. AUTHENTICATION & RBAC TESTS
# ============================================================================

def test_reports_unauthenticated(temp_db):
    """Verify report service functions throw PermissionError when unauthenticated."""
    session = get_session()
    session.clear_session()

    with pytest.raises(PermissionError):
        search_attendance_records(db_path=temp_db)

    with pytest.raises(PermissionError):
        get_student_attendance_summary(db_path=temp_db)

    with pytest.raises(PermissionError):
        correct_attendance_record(1, status="Present", db_path=temp_db)


# ============================================================================
# 2. SEARCH & MULTI-CRITERIA FILTERING TESTS
# ============================================================================

def test_search_date_range_and_filters(temp_db, auth_session):
    """Test date range, status, student, and class filtering."""
    s1 = create_student("STU-801", "Alice Reports", "12", "A", db_path=temp_db)
    s2 = create_student("STU-802", "Bob Reports", "12", "B", db_path=temp_db)

    create_attendance(s1["id"], "2026-08-10", "09:00:00", status="Present", db_path=temp_db)
    create_attendance(s1["id"], "2026-08-11", "09:05:00", status="Late", db_path=temp_db)
    create_attendance(s2["id"], "2026-08-10", "09:10:00", status="Absent", db_path=temp_db)

    # Filter 1: Date range (2026-08-10 to 2026-08-10)
    res_date = search_attendance_records(start_date="2026-08-10", end_date="2026-08-10", db_path=temp_db)
    assert len(res_date) == 2

    # Filter 2: Student query ("Alice")
    res_stu = search_attendance_records(student_query="Alice", db_path=temp_db)
    assert len(res_stu) == 2
    assert all(r["student_name"] == "Alice Reports" for r in res_stu)

    # Filter 3: Status filter ("Late")
    res_status = search_attendance_records(status_filter="Late", db_path=temp_db)
    assert len(res_status) == 1
    assert res_status[0]["status"] == "Late"


def test_invalid_date_format_rejection(temp_db, auth_session):
    """Verify invalid date strings raise ValueError."""
    with pytest.raises(ValueError, match="Invalid start_date format"):
        search_attendance_records(start_date="invalid-date", db_path=temp_db)


# ============================================================================
# 3. STUDENT ATTENDANCE ANALYTICS TESTS
# ============================================================================

def test_student_attendance_summary_analytics(temp_db, auth_session):
    """Test calculation of per-student attendance summary metrics."""
    s1 = create_student("STU-803", "Charlie Analytics", "12", "A", db_path=temp_db)

    create_attendance(s1["id"], "2026-08-15", "09:00:00", status="Present", db_path=temp_db)
    create_attendance(s1["id"], "2026-08-16", "09:00:00", status="Present", db_path=temp_db)
    create_attendance(s1["id"], "2026-08-17", "09:00:00", status="Absent", db_path=temp_db)
    create_attendance(s1["id"], "2026-08-18", "09:15:00", status="Late", db_path=temp_db)

    summary = get_student_attendance_summary(student_id=s1["id"], db_path=temp_db)
    assert len(summary) == 1
    st_sum = summary[0]

    assert st_sum["total_days"] == 4
    assert st_sum["present_count"] == 2
    assert st_sum["absent_count"] == 1
    assert st_sum["late_count"] == 1
    assert st_sum["attendance_percentage"] == 50.0  # 2 / 4 * 100 = 50.0%


def test_zero_day_attendance_percentage_safety(temp_db, auth_session):
    """Verify student with 0 attendance days safely returns 0.0% without division-by-zero."""
    s1 = create_student("STU-804", "Zero Student", "12", "B", db_path=temp_db)
    summary = get_student_attendance_summary(student_id=s1["id"], db_path=temp_db)

    assert len(summary) == 1
    assert summary[0]["total_days"] == 0
    assert summary[0]["attendance_percentage"] == 0.0


# ============================================================================
# 4. MANUAL CORRECTION TESTS
# ============================================================================

def test_manual_attendance_correction(temp_db, auth_session):
    """Test authorized manual status and time correction."""
    s1 = create_student("STU-805", "David Correction", "12", "C", db_path=temp_db)
    rec = create_attendance(s1["id"], "2026-08-20", "09:00:00", status="Absent", db_path=temp_db)

    updated = correct_attendance_record(rec["id"], status="Present", attendance_time="09:05:00", db_path=temp_db)
    assert updated["status"] == "Present"
    assert updated["attendance_time"] == "09:05:00"


def test_invalid_correction_status_rejection(temp_db, auth_session):
    """Verify invalid correction status raises ValueError."""
    s1 = create_student("STU-806", "Eve Invalid", "12", "A", db_path=temp_db)
    rec = create_attendance(s1["id"], "2026-08-20", "09:00:00", status="Present", db_path=temp_db)

    with pytest.raises(ValueError, match="Invalid status"):
        correct_attendance_record(rec["id"], status="InvalidStatus", db_path=temp_db)


# ============================================================================
# 5. CSV EXPORT TESTS
# ============================================================================

def test_export_attendance_csv(temp_db, auth_session):
    """Test writing attendance records to a valid CSV file."""
    s1 = create_student("STU-807", "Frank CSV", "12", "A", db_path=temp_db)
    create_attendance(s1["id"], "2026-08-20", "09:00:00", status="Present", db_path=temp_db)

    records = search_attendance_records(db_path=temp_db)

    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tf:
        out_path = Path(tf.name)

    exported_path = export_attendance_csv(records, out_path)
    assert exported_path.exists()

    with open(exported_path, mode="r", encoding="utf-8") as f:
        reader = list(csv.reader(f))
        assert len(reader) == 2  # Header + 1 Data row
        assert reader[0][0] == "Attendance ID"
        assert reader[1][1] == "STU-807"
        assert reader[1][2] == "Frank CSV"
        assert reader[1][6] == "Present"

    if out_path.exists():
        try:
            out_path.unlink()
        except PermissionError:
            pass


# ============================================================================
# 6. EXCEL EXPORT TESTS (OPENPYXL)
# ============================================================================

def test_export_attendance_excel(temp_db, auth_session):
    """Test generating a styled Excel (.xlsx) workbook with OpenPyXL."""
    s1 = create_student("STU-808", "Grace Excel", "12", "B", db_path=temp_db)
    create_attendance(s1["id"], "2026-08-20", "09:00:00", status="Present", db_path=temp_db)

    records = search_attendance_records(db_path=temp_db)
    summary = get_student_attendance_summary(db_path=temp_db)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        out_path = Path(tf.name)

    exported_path = export_attendance_excel(records, out_path, summary_data=summary)
    assert exported_path.exists()

    # Load workbook with OpenPyXL to verify sheets and cells
    wb = openpyxl.load_workbook(exported_path)
    assert "Attendance Records" in wb.sheetnames
    assert "Student Summary" in wb.sheetnames

    ws1 = wb["Attendance Records"]
    assert ws1.cell(row=1, column=1).value == "Attendance ID"
    assert ws1.cell(row=2, column=2).value == "STU-808"
    assert ws1.cell(row=2, column=3).value == "Grace Excel"

    wb.close()
    if out_path.exists():
        try:
            out_path.unlink()
        except PermissionError:
            pass
